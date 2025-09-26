import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import requests
import os
import shutil
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import Counter
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Spacer, Paragraph, Image
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import logging
import time
import numpy as np

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Hardcoded Excel file path and ward number
EXCEL_FILE = "VOTER_REG_STATUS\BLOUBERG WARD 16 MEMBERSHIP.xlsx"  # Replace with the actual path to your Excel file
WARD_NUMBER = 93501016

# Register Arial Black font with fallback
try:
    pdfmetrics.registerFont(TTFont('Arial-Black', 'arialbd.ttf'))
except Exception as e:
    logging.warning(f"Failed to register Arial Black font: {str(e)}. Falling back to Helvetica-Bold.")
    pdfmetrics.registerFont(TTFont('Arial-Black', 'Helvetica-Bold'))

def get_access_token(username='IECWebAPIPartyEFF', password='85316416dc5b498586ed519e670931e9'):
    """Obtain access token with retry mechanism."""
    url = 'https://api.elections.org.za/token'
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}
    body = {'grant_type': 'password', 'username': username, 'password': password}
    max_retries = 3
    retry_delay = 2

    for attempt in range(max_retries):
        try:
            response = requests.post(url, headers=headers, data=body, timeout=10)
            response.raise_for_status()
            return response.json()['access_token']
        except requests.exceptions.HTTPError as e:
            if response.status_code == 401:
                logging.error(f"Invalid credentials: {response.text}")
                if attempt == max_retries - 1:
                    raise Exception("Invalid credentials after retries.")
                time.sleep(retry_delay)
            elif response.status_code == 429:
                logging.error(f"Rate limit exceeded: {response.text}")
                time.sleep(retry_delay * (attempt + 1))
            else:
                logging.error(f"HTTP error: {response.status_code} - {response.text}")
                raise
        except requests.exceptions.RequestException as e:
            logging.error(f"Failed to obtain access token: {str(e)}")
            if attempt == max_retries - 1:
                raise Exception(f"Failed to obtain access token after {max_retries} attempts: {str(e)}")
            time.sleep(retry_delay)
    raise Exception("Failed to obtain access token.")

def fetch_voter_data(id_number, index, access_token):
    """Fetch voter data from API."""
    base_url = 'https://api.elections.org.za/'
    endpoint = "api/Voters/IDNumber/"
    headers = {'Content-Type': 'application/json', 'Authorization': f'bearer {access_token}'}
    try:
        url = f'{base_url}{endpoint}{id_number}'
        response = requests.get(url, headers=headers, timeout=10)
        if response.status_code == 200:
            data = response.json()
            logging.info(f"Processed ID {id_number} at index {index}")
            return {
                'index': index,
                'id': id_number,
                'bRegistered': data.get('bRegistered', False),
                'ward_id': data.get('VotingStation', {}).get('Delimitation', {}).get('WardID', None),
                'province': data.get('VotingStation', {}).get('Delimitation', {}).get('Province', '') or '',
                'municipality': data.get('VotingStation', {}).get('Delimitation', {}).get('Municipality', '') or '',
                'voting_station': data.get('VotingStation', {}).get('Name', '') or '',
                'vd_number': data.get('VotingStation', {}).get('Delimitation', {}).get('VDNumber', '') or '',
                'suburb': data.get('VotingStation', {}).get('Location', {}).get('Suburb', '') or '',
                'street': data.get('VotingStation', {}).get('Location', {}).get('Street', '') or ''
            }
        else:
            logging.error(f"Failed to retrieve data for ID {id_number}. Status code: {response.status_code}")
            return None
    except Exception as e:
        logging.error(f"Error processing ID {id_number}: {str(e)}")
        return None

def copy_row_to_sheet(sheet, index, source_ws, ward_number=None):
    """Copy all columns from source worksheet to target sheet for the given row."""
    max_col = source_ws.max_column
    for col in range(1, max_col + 1):
        value = source_ws.cell(row=index + 2, column=col).value
        if col == 8:  # ID Number column
            value = str(value).rjust(13, '0') if value else ""
        elif col == 5 and sheet.title == 'RegisteredInWard':
            value = ward_number if source_ws.cell(row=index + 2, column=col).value is None else source_ws.cell(row=index + 2, column=col).value
        if value is None or (isinstance(value, float) and np.isnan(value)):
            value = ""
        sheet.cell(row=index + 2, column=col).value = value

def remove_empty_rows(sheet):
    """Remove rows where all cells are empty, excluding metadata row."""
    if sheet.max_row < 2:
        return 0
    del_rows = []
    for row in range(2, sheet.max_row + 1):
        is_empty = all(sheet.cell(row=row, column=col).value is None or str(sheet.cell(row=row, column=col).value).strip() == "" for col in range(1, sheet.max_column + 1))
        if is_empty:
            del_rows.append(row)
    for row in reversed(del_rows):
        sheet.delete_rows(row)
    return len(del_rows)

def sort_sheet_by_voting_station(sheet, voting_station_col=4):
    """Sort rows by voting station, preserving metadata row."""
    if sheet.max_row <= 2:
        return
    data = []
    for row in range(2, sheet.max_row + 1):
        row_data = [sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)]
        if len(row_data) >= 8 and row_data[7] is not None:
            row_data[7] = str(row_data[7]).rjust(13, '0')
        row_data = ["" if x is None or (isinstance(x, float) and np.isnan(x)) else x for x in row_data]
        data.append((row_data[voting_station_col - 1] or '', row_data))
    data.sort(key=lambda x: x[0].lower())
    for row in range(2, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=col).value = None
    for row_idx, (_, row_data) in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx).value = value

def generate_pdf_report(ward_data, not_in_ward_data, ward_number, output_path, municipality):
    """Generate a PDF report for the given ward."""
    doc = SimpleDocTemplate(
        output_path,
        pagesize=landscape(letter),
        leftMargin=0.5*inch,
        rightMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.75*inch
    )
    elements = []
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title_style.fontName = 'Helvetica-Bold'
    title_style.fontSize = 16
    title_style.leading = 18
    title_style.alignment = 1

    header_style = styles['Normal']
    header_style.fontName = 'Helvetica'
    header_style.fontSize = 10
    header_style.leading = 12
    header_style.alignment = 0

    logo_path = "RedLogoSquare.png"
    logo_width = 68
    logo_height = 68
    logo_image = None
    if os.path.exists(logo_path):
        try:
            logo_image = Image(logo_path, width=logo_width, height=logo_height)
            logo_image.hAlign = 'CENTER'
            logo_image.vAlign = 'TOP'
        except Exception as e:
            logging.error(f"Failed to load logo: {str(e)}")

    title = Paragraph("FORM A: ATTENDANCE REGISTER", title_style)
    total_voters = len(ward_data) + len(not_in_ward_data)
    quorum = total_voters // 2 + 1
    total_voting_stations = len(set(row.get("REGISTERED VD", "") for row in ward_data if row.get("REGISTERED VD", "")))
    province = next((row.get("PROVINCE", "Unknown") for row in ward_data if row.get("PROVINCE", "Unknown") != "Unknown"), "Unknown")

    left_header_data = [
        [f"PROVINCE: {province}"],
        [f"TOTAL MEMBERSHIP IN GOOD STANDING: {total_voters}"],
        [f"QUORUM: {quorum}"],
        ["DATE OF BPA/BGA:"]
    ]
    right_header_data = [
        [f"SUB REGION: {municipality}"],
        [f"WARD: {ward_number}"],
        ["BPA: |_| BGA: |_|"],
        [f"TOTAL NUMBER OF VOTING STATIONS: {total_voting_stations}"]
    ]

    font_name = 'Helvetica'
    font_size = 10
    padding = 6
    left_width = max(pdfmetrics.stringWidth(cell[0], font_name, font_size) + 2 * padding for cell in left_header_data)
    right_width = max(pdfmetrics.stringWidth(cell[0], font_name, font_size) + 2 * padding for cell in right_header_data)

    left_header_table = Table(left_header_data, colWidths=[left_width])
    left_header_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('LEADING', (0, 0), (-1, -1), 12),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0, colors.white),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))

    right_header_table = Table(right_header_data, colWidths=[right_width])
    right_header_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('LEADING', (0, 0), (-1, -1), 12),
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0, colors.white),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))

    elements.append(Table([[title]], colWidths=[683], rowHeights=[50], style=[
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (0, 0), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0, colors.white),
    ]))

    elements.append(Spacer(1, 6))
    separator_line = Table([[""]], colWidths=[683], rowHeights=[2])
    separator_line.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.black),
        ('GRID', (0, 0), (-1, -1), 0, colors.transparent),
    ]))
    elements.append(separator_line)
    elements.append(Spacer(1, 6))

    if logo_image:
        logo_table = Table([[logo_image]], colWidths=[683], rowHeights=[logo_height])
        logo_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0, colors.white),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(logo_table)
        elements.append(Spacer(1, 2))

    max_page_width = 683
    middle_width = max_page_width - left_width - right_width
    top_frame_data = [[left_header_table, "", right_header_table]]
    top_frame_table = Table(top_frame_data, colWidths=[left_width, middle_width, right_width], rowHeights=[None])
    top_frame_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0, colors.white),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(top_frame_table)
    elements.append(Spacer(1, 8))

    headers = ["NUM...", "NAME", "WARD NUMBER", "ID NUMBER", "CELL NUMBER", "REGISTERED VD", "SIGNATURE", "NEW CELL NUM"]
    grouped_data = {}
    for row in ward_data:
        voting_station = str(row.get("REGISTERED VD", ""))
        if voting_station not in grouped_data:
            grouped_data[voting_station] = []
        grouped_data[voting_station].append(row)

    sorted_voting_stations = sorted(grouped_data.keys(), key=lambda x: x.lower() if x else "")
    table_data = [headers]
    all_text_for_width = [headers]
    group_header_rows = []
    voter_count_rows = []
    current_row = 1

    for voting_station in sorted_voting_stations:
        vd_number = grouped_data[voting_station][0].get("VD NUMBER", "") if grouped_data[voting_station] else ""
        group_header = [f"Voting Station: {voting_station or 'Unknown'} (VDNumber: {vd_number})", "", "", "", "", "", "", ""]
        table_data.append(group_header)
        all_text_for_width.append(group_header)
        group_header_rows.append(current_row)
        current_row += 1

        for i, row in enumerate(grouped_data[voting_station], start=1):
            cell_number = str(row.get("CELL NUMBER", "")).replace(".0", "")
            new_cell_num = str(row.get("NEW CELL NUM", "")).replace(".0", "")
            name = str(row.get("NAME", ""))
            row_data = [
                str(i),
                name,
                str(row.get("WARD NUMBER", "")),
                str(row.get("ID NUMBER", "")),
                cell_number,
                str(row.get("REGISTERED VD", "")),
                str(row.get("SIGNATURE", "")),
                new_cell_num
            ]
            table_data.append(row_data)
            all_text_for_width.append(row_data)
            current_row += 1

        voter_count = len(grouped_data[voting_station])
        voter_count_row = [f"Total Voters in {voting_station or 'Unknown'}: {voter_count}", "", "", "", "", "", "", ""]
        table_data.append(voter_count_row)
        all_text_for_width.append(voter_count_row)
        voter_count_rows.append(current_row)
        current_row += 1

    font_name = 'Helvetica'
    font_name_bold = 'Helvetica-Bold'
    font_size = 8
    max_page_width = 683
    padding = 8
    max_widths = [50, 160, 80, 110, 90, 120, 80, 80]
    min_widths = [50, 100, 80, 110, 60, 120, 80, 80]
    col_max_widths = [0] * len(headers)

    for row in all_text_for_width:
        for idx, cell in enumerate(row):
            text = str(cell)
            font = 'Arial-Black' if (idx == 0 and row[0].startswith("Voting Station:")) else \
                   font_name_bold if (idx == 0 and row[0].startswith("Total Voters in") or row == headers) else font_name
            text_width = pdfmetrics.stringWidth(text, font, font_size) + padding
            col_max_widths[idx] = max(col_max_widths[idx], text_width)

    for idx in range(len(headers)):
        col_max_widths[idx] = min(max_widths[idx], max(min_widths[idx], col_max_widths[idx]))

    priority_columns = [1, 5]
    total_width = sum(col_max_widths)
    if total_width > max_page_width:
        excess_width = total_width - max_page_width
        non_priority_columns = [i for i in range(len(headers)) if i not in priority_columns]
        total_non_priority_width = sum(col_max_widths[i] for i in non_priority_columns)
        if total_non_priority_width > 0:
            scale_factor = (total_non_priority_width - excess_width) / total_non_priority_width
            if scale_factor > 0:
                for i in non_priority_columns:
                    col_max_widths[i] *= scale_factor
            else:
                scale_factor = max_page_width / total_width
                col_max_widths = [w * scale_factor for w in col_max_widths]
    elif total_width < max_page_width:
        extra_width = max_page_width - total_width
        extra_per_priority = extra_width / len(priority_columns)
        for i in priority_columns:
            col_max_widths[i] = min(col_max_widths[i] + extra_per_priority, max_widths[i])

    total_width = sum(col_max_widths)
    if abs(total_width - max_page_width) > 0.01:
        scale_factor = max_page_width / total_width
        col_max_widths = [w * scale_factor for w in col_max_widths]

    cell_style = styles['Normal']
    cell_style.fontName = font_name
    cell_style.fontSize = font_size
    cell_style.leading = font_size * 1.2
    cell_style.alignment = 0

    header_style = styles['Normal']
    header_style.fontName = font_name_bold
    header_style.fontSize = font_size
    header_style.leading = font_size * 1.2
    header_style.alignment = 0

    group_header_style = styles['Normal']
    group_header_style.fontName = 'Arial-Black'
    group_header_style.fontSize = font_size
    group_header_style.leading = font_size * 1.2
    group_header_style.alignment = 0

    voter_count_style = styles['Normal']
    voter_count_style.fontName = font_name_bold
    voter_count_style.fontSize = font_size
    voter_count_style.leading = font_size * 1.2
    voter_count_style.alignment = 0

    wrapped_table_data = []
    for row_idx, row in enumerate(table_data):
        wrapped_row = []
        for col_idx, cell in enumerate(row):
            text = str(cell)
            style = (header_style if row_idx == 0 else
                     group_header_style if row_idx in group_header_rows else
                     voter_count_style if row_idx in voter_count_rows else
                     cell_style)
            p = Paragraph(text, style)
            wrapped_row.append(p)
        wrapped_table_data.append(wrapped_row)

    table = Table(wrapped_table_data, colWidths=col_max_widths, rowHeights=None)
    style_commands = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTSIZE', (0, 0), (-1, -1), font_size),
        ('LEADING', (0, 0), (-1, -1), font_size * 1.2),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
    ]

    for row_idx in group_header_rows:
        style_commands.extend([
            ('BACKGROUND', (0, row_idx), (-1, row_idx), colors.lightgrey),
            ('SPAN', (0, row_idx), (-1, row_idx)),
            ('VALIGN', (0, row_idx), (-1, -1), 'TOP'),
        ])

    for row_idx in voter_count_rows:
        style_commands.extend([
            ('BACKGROUND', (0, row_idx), (-1, row_idx), colors.yellow),
            ('SPAN', (0, row_idx), (-1, row_idx)),
            ('VALIGN', (0, row_idx), (-1, -1), 'TOP'),
        ])

    table.setStyle(TableStyle(style_commands))
    elements.append(table)

    if not_in_ward_data:
        elements.append(Spacer(1, 12))
        not_in_ward_title = Paragraph("Not Registered In Ward Data", group_header_style)
        not_in_ward_title_table = Table([[not_in_ward_title]], colWidths=[max_page_width], rowHeights=[None])
        not_in_ward_title_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(not_in_ward_title_table)
        elements.append(Spacer(1, 4))

        not_in_ward_table_data = [headers]
        not_in_ward_all_text_for_width = [headers]
        not_in_ward_header_row = [0]

        for i, row in enumerate(not_in_ward_data, start=1):
            cell_number = str(row.get("CELL NUMBER", "")).replace(".0", "")
            new_cell_num = str(row.get("NEW CELL NUM", "")).replace(".0", "")
            name = str(row.get("NAME", ""))
            row_data = [
                str(i),
                name,
                str(row.get("WARD NUMBER", "")),
                str(row.get("ID NUMBER", "")),
                cell_number,
                str(row.get("REGISTERED VD", "")),
                str(row.get("SIGNATURE", "")),
                new_cell_num
            ]
            not_in_ward_table_data.append(row_data)
            not_in_ward_all_text_for_width.append(row_data)

        total_not_in_ward = len(not_in_ward_data)
        total_row = [f"Total Not Registered in Ward Voters: {total_not_in_ward}", "", "", "", "", "", "", ""]
        not_in_ward_table_data.append(total_row)
        not_in_ward_all_text_for_width.append(total_row)
        not_in_ward_voter_count_row = len(not_in_ward_table_data) - 1

        not_in_ward_col_max_widths = [0] * len(headers)
        for row in not_in_ward_all_text_for_width:
            for idx, cell in enumerate(row):
                text = str(cell)
                font = font_name_bold if (idx == 0 and (row[0].startswith("Total Not Registered in Ward Voters") or row == headers)) else font_name
                text_width = pdfmetrics.stringWidth(text, font, font_size) + padding
                not_in_ward_col_max_widths[idx] = max(not_in_ward_col_max_widths[idx], text_width)

        for idx in range(len(headers)):
            not_in_ward_col_max_widths[idx] = min(max_widths[idx], max(min_widths[idx], not_in_ward_col_max_widths[idx]))

        total_width = sum(not_in_ward_col_max_widths)
        if total_width > max_page_width:
            excess_width = total_width - max_page_width
            non_priority_columns = [i for i in range(len(headers)) if i not in priority_columns]
            total_non_priority_width = sum(not_in_ward_col_max_widths[i] for i in non_priority_columns)
            if total_non_priority_width > 0:
                scale_factor = (total_non_priority_width - excess_width) / total_non_priority_width
                if scale_factor > 0:
                    for i in non_priority_columns:
                        not_in_ward_col_max_widths[i] *= scale_factor
                else:
                    scale_factor = max_page_width / total_width
                    not_in_ward_col_max_widths = [w * scale_factor for w in not_in_ward_col_max_widths]
        elif total_width < max_page_width:
            extra_width = max_page_width - total_width
            extra_per_priority = extra_width / len(priority_columns)
            for i in priority_columns:
                not_in_ward_col_max_widths[i] = min(not_in_ward_col_max_widths[i] + extra_per_priority, max_widths[i])

        total_width = sum(not_in_ward_col_max_widths)
        if abs(total_width - max_page_width) > 0.01:
            scale_factor = max_page_width / total_width
            not_in_ward_col_max_widths = [w * scale_factor for w in not_in_ward_col_max_widths]

        wrapped_not_in_ward_table_data = []
        for row_idx, row in enumerate(not_in_ward_table_data):
            wrapped_row = []
            for col_idx, cell in enumerate(row):
                text = str(cell)
                style = (header_style if row_idx in not_in_ward_header_row else
                         voter_count_style if row_idx == not_in_ward_voter_count_row else
                         cell_style)
                p = Paragraph(text, style)
                wrapped_row.append(p)
            wrapped_not_in_ward_table_data.append(wrapped_row)

        not_in_ward_table = Table(wrapped_not_in_ward_table_data, colWidths=not_in_ward_col_max_widths, rowHeights=None)
        not_in_ward_style_commands = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTSIZE', (0, 0), (-1, -1), font_size),
            ('LEADING', (0, 0), (-1, -1), font_size * 1.2),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, not_in_ward_voter_count_row), (-1, not_in_ward_voter_count_row), colors.yellow),
            ('SPAN', (0, not_in_ward_voter_count_row), (-1, not_in_ward_voter_count_row)),
            ('VALIGN', (0, not_in_ward_voter_count_row), (-1, -1), 'TOP'),
        ]

        not_in_ward_table.setStyle(TableStyle(not_in_ward_style_commands))
        elements.append(not_in_ward_table)

    def on_page(canvas, doc):
        canvas.saveState()
        footer_style = styles['Normal']
        footer_style.fontName = 'Helvetica'
        footer_style.fontSize = 10
        footer_style.leading = 12
        footer_style.alignment = 0
        sub_region_text = f"SUB REGION: {municipality}"
        canvas.setFont('Helvetica', 10)
        canvas.setFillColor(colors.black)
        canvas.drawString(doc.leftMargin, doc.bottomMargin - 10, sub_region_text)
        ward_text = f"WARD: {ward_number}"
        ward_text_width = pdfmetrics.stringWidth(ward_text, 'Helvetica', 10)
        canvas.drawString(doc.leftMargin + doc.width - ward_text_width, doc.bottomMargin - 10, ward_text)
        page_number_text = f"Page {canvas.getPageNumber()}"
        canvas.drawCentredString(doc.leftMargin + doc.width / 2, doc.bottomMargin - 10, page_number_text)
        canvas.restoreState()

    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)
    logging.info(f"Generated PDF report: {output_path}")

def process_excel_file(excel_file, ward_number):
    """Process a single Excel file and generate output files and PDF report."""
    column_index = 8  # ID Number column
    output_dir = "VOTER_REG_STATUS/ProcessedFiles"
    os.makedirs(output_dir, exist_ok=True)
    start_time = time.time()

    counter_false = 0
    counter_true = 0
    deceased_counter = 0
    not_in_ward_counter = 0
    registered_in_ward_data = []
    not_registered_in_ward_data = []
    voting_station_counts = Counter()

    if not os.path.exists(excel_file):
        logging.error(f"Excel file {excel_file} does not exist.")
        print(f"Excel file {excel_file} does not exist.")
        return

    try:
        access_token = get_access_token()
    except Exception as e:
        logging.error(f"Failed to obtain access token: {str(e)}")
        print(f"Error obtaining access token: {str(e)}")
        return

    excel_input_file = os.path.basename(excel_file)
    output_file_path = os.path.join(output_dir, excel_input_file)
    try:
        shutil.copy(excel_file, output_file_path)
        logging.info(f"Copied file to: {output_file_path}")
        print(f"Copied file to: {output_file_path}")
    except (shutil.Error, OSError) as e:
        logging.error(f"Error copying file {excel_input_file}: {str(e)}")
        print(f"Error copying file {excel_input_file}: {str(e)}")
        return

    try:
        wb = openpyxl.load_workbook(output_file_path)
        ws = wb.active
    except Exception as e:
        logging.error(f"Error loading copied file {output_file_path}: {str(e)}")
        print(f"Error loading copied file {output_file_path}: {str(e)}")
        return

    df = pd.read_excel(excel_file, header=0)
    df = df.fillna("")
    column_mapping = {
        'Firstname': 'Firstname',
        'Surname': 'Surname',
        'Ward Number': 'Ward Number',
        'ID Number': 'ID Number',
        'Cell Number': 'Cell Number'
    }
    df = df.rename(columns=column_mapping)

    for sheet_name in ['NotRegisteredVoter', 'RegisteredInWard', 'Deceased', 'NotRegisteredInWard']:
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)

    ws1 = wb['NotRegisteredVoter']
    ws2 = wb['RegisteredInWard']
    ws3 = wb['Deceased']
    ws4 = wb['NotRegisteredInWard']

    if ws.max_row >= 1:
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            for sheet in [ws1, ws2, ws3, ws4]:
                sheet.cell(row=1, column=col).value = value

    for sheet in [ws, ws1, ws2, ws3, ws4]:
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column_index)].width = 15
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=column_index)
            cell.number_format = '0000000000000'

    id_list = []
    for j in range(2, ws.max_row + 1):
        if ws.cell(row=j, column=column_index).value is not None:
            id_number = str(ws.cell(row=j, column=column_index).value).rjust(13, '0')
            id_list.append((id_number, j - 2))
            ws.cell(row=j, column=column_index).value = id_number

    results = []
    with ThreadPoolExecutor(max_workers=10) as executor:
        future_to_index = {executor.submit(fetch_voter_data, id_num, i, access_token): i for id_num, i in id_list}
        for future in as_completed(future_to_index):
            result = future.result()
            if result:
                results.append(result)

    for result in results:
        i = result['index']
        row = i + 2
        address = f"{result['suburb']} {result['street']}".strip()
        ward_number_to_use = result['ward_id'] if result['ward_id'] is not None else ward_number
        ws.cell(row=row, column=4).value = result['voting_station']
        ws.cell(row=row, column=5).value = ward_number_to_use
        ws.cell(row=row, column=1).value = result['province']
        ws.cell(row=row, column=3).value = result['municipality']
        if ws.cell(row=row, column=14).value in ["N/A", "CENTURION", None]:
            ws.cell(row=row, column=14).value = address

        try:
            excel_row = df.iloc[i]
            firstname = str(excel_row.get("Firstname", ""))
            surname = str(excel_row.get("Surname", ""))
            name = f"{firstname} {surname}".strip().upper()
            cell_number = str(excel_row.get("Cell Number", "")).replace(".0", "")
            voter_record = {
                "NAME": name if name else "",
                "WARD NUMBER": str(ward_number_to_use),
                "ID NUMBER": str(result['id']),
                "CELL NUMBER": cell_number if cell_number else "",
                "REGISTERED VD": str(result['voting_station']),
                "VD NUMBER": str(result['vd_number']),
                "SIGNATURE": "",
                "NEW CELL NUM": "",
                "PROVINCE": str(result['province']),
                "MUNICIPALITY": str(result['municipality'])
            }
        except IndexError:
            voter_record = {
                "NAME": "",
                "WARD NUMBER": str(ward_number_to_use),
                "ID NUMBER": str(result['id']),
                "CELL NUMBER": "",
                "REGISTERED VD": str(result['voting_station']),
                "VD NUMBER": str(result['vd_number']),
                "SIGNATURE": "",
                "NEW CELL NUM": "",
                "PROVINCE": str(result['province']),
                "MUNICIPALITY": str(result['municipality'])
            }

        if result['voting_station'] and not result['bRegistered']:
            deceased_counter += 1
            ws.cell(row=row, column=26).value = "DECEASED"
            ws.cell(row=row, column=26).fill = PatternFill(fgColor='ffff00', fill_type='solid')
            ws.cell(row=row, column=8).fill = PatternFill(fgColor='ffff00', fill_type='solid')
            ws.cell(row=row, column=5).fill = PatternFill(fgColor='ffff00', fill_type='solid')
            copy_row_to_sheet(ws3, i, ws, ward_number)
        elif result['bRegistered'] and str(result['ward_id']) == str(ward_number):
            counter_true += 1
            ws.cell(row=row, column=26).value = "Registered In Ward"
            ws.cell(row=row, column=26).fill = PatternFill(fgColor='00ff00', fill_type='solid')
            ws.cell(row=row, column=5).fill = PatternFill(fgColor='00ff00', fill_type='solid')
            registered_in_ward_data.append(voter_record)
            copy_row_to_sheet(ws2, i, ws, ward_number)
            voting_station_counts[result['voting_station']] += 1
        elif result['bRegistered'] and str(result['ward_id']) != str(ward_number):
            not_in_ward_counter += 1
            ws.cell(row=row, column=26).value = "Not Registered In Ward"
            ws.cell(row=row, column=26).fill = PatternFill(fgColor='0000FF', fill_type='solid')
            ws.cell(row=row, column=5).fill = PatternFill(fgColor='0000FF', fill_type='solid')
            not_registered_in_ward_data.append(voter_record)
            copy_row_to_sheet(ws4, i, ws, ward_number)
        else:
            counter_false += 1
            ws.cell(row=row, column=26).value = "Not Registered Voter"
            ws.cell(row=row, column=26).fill = PatternFill(fgColor='ff0000', fill_type='solid')
            ws.cell(row=row, column=5).fill = PatternFill(fgColor='ff0000', fill_type='solid')
            copy_row_to_sheet(ws1, i, ws, ward_number)

    remove_empty_rows(ws1)
    remove_empty_rows(ws2)
    remove_empty_rows(ws3)
    remove_empty_rows(ws4)

    sort_sheet_by_voting_station(ws2)
    sort_sheet_by_voting_station(ws4)

    try:
        wb.save(output_file_path)
        logging.info(f"Saved processed file: {output_file_path}")
        print(f"Saved processed file: {output_file_path}")
    except (PermissionError, OSError) as e:
        logging.error(f"Error saving {output_file_path}: {str(e)}")
        print(f"Error saving {output_file_path}: {str(e)}")
        return

    total_processed = counter_true + counter_false + deceased_counter + not_in_ward_counter
    print(f"\nResults for {excel_input_file} (Ward {ward_number}):")
    print(f"Total IDs Processed: {total_processed}")
    print(f"Registered in Ward: {counter_true}")
    print(f"Not Registered in Ward: {not_in_ward_counter}")
    print(f"Not Registered Voters: {counter_false}")
    print(f"Deceased: {deceased_counter}")
    if voting_station_counts:
        print("Voting Station Counts:")
        for vs, count in voting_station_counts.items():
            print(f"  {vs}: {count} voters")

    end_time = time.time()
    print(f"\nProcessing completed in {end_time - start_time:.2f} seconds.")

    # Generate PDF report
    output_pdf = os.path.join(output_dir, f"Ward_{ward_number}_Attendance_Register.pdf")
    municipality = next((row.get("MUNICIPALITY", "Unknown") for row in registered_in_ward_data if row.get("MUNICIPALITY", "Unknown") != "Unknown"), "Unknown")
    generate_pdf_report(registered_in_ward_data, not_registered_in_ward_data, ward_number, output_pdf, municipality)
    print(f"Generated PDF report: {output_pdf}")

if __name__ == "__main__":
    process_excel_file(EXCEL_FILE, WARD_NUMBER)