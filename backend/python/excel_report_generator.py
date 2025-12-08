"""
Excel Report Generator for Bulk Upload Processing
Generates comprehensive Excel reports with multiple sheets containing detailed statistics
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List, Optional
import os
import logging

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Generate comprehensive Excel reports for bulk upload processing"""
    
    # Color schemes
    COLORS = {
        'header': 'D3D3D3',      # Light gray
        'success': 'C6EFCE',     # Light green
        'error': 'FFC7CE',       # Light red
        'warning': 'FFEB9C',     # Light yellow
        'info': 'BDD7EE'         # Light blue
    }
    
    def __init__(self, original_filename: str, output_dir: str):
        """
        Initialize report generator
        
        Args:
            original_filename: Original uploaded file name
            output_dir: Directory to save the report
        """
        self.original_filename = original_filename
        self.output_dir = output_dir
        self.timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        
        # Generate report filename
        base_name = os.path.splitext(original_filename)[0]
        self.report_filename = f"{base_name}_REPORT_{self.timestamp}.xlsx"
        self.report_path = os.path.join(output_dir, self.report_filename)
        
    def generate_report(self,
                       df_original: pd.DataFrame,
                       df_verified: pd.DataFrame,
                       processing_stats: Dict,
                       invalid_ids: List[Dict],
                       duplicates: List[Dict],
                       different_ward: List[Dict],
                       not_registered: List[Dict],
                       successfully_imported: List[Dict],
                       existing_members: Optional[List[Dict]] = None) -> str:
        """
        Generate comprehensive Excel report with multiple sheets

        Args:
            df_original: Original uploaded DataFrame
            df_verified: Verified DataFrame with IEC data
            processing_stats: Dictionary with processing statistics
            invalid_ids: List of invalid ID records
            duplicates: List of duplicate records
            different_ward: List of members in different ward
            not_registered: List of members not registered
            successfully_imported: List of successfully imported records
            existing_members: List of members that already existed in database (updated)

        Returns:
            Path to generated report file
        """
        logger.info(f"📊 Generating Excel report: {self.report_filename}")

        # Create Excel writer
        with pd.ExcelWriter(self.report_path, engine='openpyxl') as writer:
            # Sheet 1: Summary
            self._create_summary_sheet(writer, processing_stats)

            # Sheet 2: All Uploaded Rows (Original Data)
            self._create_all_uploaded_rows_sheet(writer, df_original)

            # Sheet 3: Invalid IDs
            self._create_invalid_ids_sheet(writer, invalid_ids)

            # Sheet 4: Duplicates
            self._create_duplicates_sheet(writer, duplicates)

            # Sheet 5: Different Ward
            self._create_different_ward_sheet(writer, different_ward)

            # Sheet 6: Not Registered
            self._create_not_registered_sheet(writer, not_registered)

            # Sheet 7: Successfully Imported (New Members)
            self._create_successfully_imported_sheet(writer, successfully_imported)

            # Sheet 8: Existing Members (Updated)
            if existing_members:
                self._create_existing_members_sheet(writer, existing_members)

        # Apply formatting
        self._apply_formatting()

        logger.info(f"✅ Report generated successfully: {self.report_path}")
        return self.report_path

    def _build_summary_values(self, stats: Dict) -> List:
        """Build summary values list dynamically based on stats"""
        values = [
            self.original_filename,
            self.timestamp,
            '',
            '',
            stats.get('total_records', 0),
            stats.get('valid_ids', 0),
            stats.get('invalid_ids', 0),
            '',
            '',
            stats.get('verified_count', 0),
            stats.get('registered_in_ward', 0),
            stats.get('different_ward', 0),
            stats.get('not_registered', 0),
            stats.get('deceased', 0),
            stats.get('api_errors', 0),
        ]

        # Add rate limit values if applicable
        if stats.get('rate_limit_hit', False):
            reset_time = stats.get('rate_limit_reset_time', 0)
            reset_str = datetime.fromtimestamp(reset_time / 1000).strftime('%Y-%m-%d %H:%M:%S') if reset_time else 'N/A'

            values.extend([
                '',
                '',
                'EXCEEDED',
                stats.get('rows_processed_before_limit', 0),
                f"{stats.get('rate_limit_message', 'Rate limit exceeded')} (Resets at: {reset_str})",
            ])

        values.extend([
            '',
            '',
            stats.get('vd_populated', 0),
            stats.get('vd_empty', 0),
            '',
            '',
            stats.get('unique_ids', 0),
            stats.get('duplicate_ids', 0),
            stats.get('duplicate_records', 0),
            '',
            '',
            stats.get('imported', 0),
            stats.get('skipped', 0),
            stats.get('new_members', 0),
            stats.get('existing_members', 0),
            stats.get('processing_time', 0),
            stats.get('processing_speed', 0),
            stats.get('status', 'Unknown')
        ])

        return values

    def _build_summary_percentages(self, stats: Dict) -> List:
        """Build summary percentages list dynamically based on stats"""
        percentages = [
            '', '', '', '', '',
            f"{(stats.get('valid_ids', 0) / stats.get('total_records', 1) * 100):.1f}%",
            f"{(stats.get('invalid_ids', 0) / stats.get('total_records', 1) * 100):.1f}%",
            '', '', '',
            f"{(stats.get('registered_in_ward', 0) / stats.get('total_records', 1) * 100):.1f}%",
            f"{(stats.get('different_ward', 0) / stats.get('total_records', 1) * 100):.1f}%",
            f"{(stats.get('not_registered', 0) / stats.get('total_records', 1) * 100):.1f}%",
            '', '',
        ]

        # Add rate limit percentages if applicable
        if stats.get('rate_limit_hit', False):
            total = stats.get('total_records', 1)
            processed = stats.get('rows_processed_before_limit', 0)
            percentages.extend([
                '', '', '',
                f"{(processed / total * 100):.1f}%",
                '',
            ])

        percentages.extend([
            '', '',
            f"{(stats.get('vd_populated', 0) / stats.get('total_records', 1) * 100):.1f}%",
            f"{(stats.get('vd_empty', 0) / stats.get('total_records', 1) * 100):.1f}%",
            '', '', '', '', '', '', '',
            f"{(stats.get('imported', 0) / stats.get('total_records', 1) * 100):.1f}%",
            '',
            f"{(stats.get('new_members', 0) / stats.get('imported', 1) * 100):.1f}%",
            f"{(stats.get('existing_members', 0) / stats.get('imported', 1) * 100):.1f}%",
            '', '', ''
        ])

        return percentages

    def _create_summary_sheet(self, writer: pd.ExcelWriter, stats: Dict):
        """Create summary sheet with overall statistics"""

        # Build metrics list dynamically
        metrics = [
            'File Name',
            'Processing Date',
            '',
            '=== OVERALL STATISTICS ===',
            'Total Records',
            'Valid ID Numbers',
            'Invalid ID Numbers',
            '',
            '=== IEC VERIFICATION ===',
            'Successfully Verified',
            'Registered in Correct Ward',
            'Registered in Different Ward',
            'Not Registered to Vote',
            'Deceased',
            'API Errors',
        ]

        # Add rate limit section if applicable
        if stats.get('rate_limit_hit', False):
            metrics.extend([
                '',
                '=== IEC API RATE LIMIT ===',
                'Rate Limit Status',
                'Records Processed Before Limit',
                'Rate Limit Message',
            ])

        metrics.extend([
            '',
            '=== VD NUMBER POPULATION ===',
            'Records with VD Numbers',
            'Records without VD Numbers',
            '',
            '=== DUPLICATE DETECTION ===',
            'Unique ID Numbers',
            'Duplicate ID Numbers',
            'Total Duplicate Records',
            '',
            '=== DATABASE INGESTION ===',
            'Records Imported',
            'Records Skipped',
            'New Members (Inserted)',
            'Existing Members (Updated)',
            'Processing Time (seconds)',
            'Processing Speed (records/sec)',
            'Status'
        ])

        summary_data = {
            'Metric': metrics,
            'Value': self._build_summary_values(stats),
            'Percentage': self._build_summary_percentages(stats)
        }
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)

    def _create_all_uploaded_rows_sheet(self, writer: pd.ExcelWriter, df_original: pd.DataFrame):
        """Create sheet with all uploaded rows from the original file"""
        if df_original is None or df_original.empty:
            df_all_rows = pd.DataFrame({'Message': ['No data uploaded']})
        else:
            # Use the original dataframe as-is to show all uploaded rows
            df_all_rows = df_original.copy()

        df_all_rows.to_excel(writer, sheet_name='All Uploaded Rows', index=False)

    def _create_invalid_ids_sheet(self, writer: pd.ExcelWriter, invalid_ids: List[Dict]):
        """Create sheet with invalid ID numbers"""
        if not invalid_ids:
            df_invalid = pd.DataFrame({'Message': ['No invalid ID numbers found']})
        else:
            df_invalid = pd.DataFrame(invalid_ids)

        df_invalid.to_excel(writer, sheet_name='Invalid IDs', index=False)

    def _create_duplicates_sheet(self, writer: pd.ExcelWriter, duplicates: List[Dict]):
        """Create sheet with duplicate records"""
        if not duplicates:
            df_duplicates = pd.DataFrame({'Message': ['No duplicate records found']})
        else:
            df_duplicates = pd.DataFrame(duplicates)

        df_duplicates.to_excel(writer, sheet_name='Duplicates', index=False)

    def _create_different_ward_sheet(self, writer: pd.ExcelWriter, different_ward: List[Dict]):
        """Create sheet with members in different ward"""
        if not different_ward:
            df_different = pd.DataFrame({'Message': ['No members registered in different ward']})
        else:
            df_different = pd.DataFrame(different_ward)

        df_different.to_excel(writer, sheet_name='Different Ward', index=False)

    def _create_not_registered_sheet(self, writer: pd.ExcelWriter, not_registered: List[Dict]):
        """Create sheet with members not registered to vote"""
        if not not_registered:
            df_not_reg = pd.DataFrame({'Message': ['No unregistered members found']})
        else:
            df_not_reg = pd.DataFrame(not_registered)

        df_not_reg.to_excel(writer, sheet_name='Not Registered', index=False)

    def _create_successfully_imported_sheet(self, writer: pd.ExcelWriter, successfully_imported: List[Dict]):
        """Create sheet with successfully imported records (new members only)"""
        if not successfully_imported:
            df_imported = pd.DataFrame({'Message': ['No new members imported']})
        else:
            df_imported = pd.DataFrame(successfully_imported)

        df_imported.to_excel(writer, sheet_name='New Members', index=False)

    def _create_existing_members_sheet(self, writer: pd.ExcelWriter, existing_members: List[Dict]):
        """Create sheet with existing members that were updated"""
        if not existing_members:
            df_existing = pd.DataFrame({'Message': ['No existing members were updated']})
        else:
            df_existing = pd.DataFrame(existing_members)

        df_existing.to_excel(writer, sheet_name='Existing Members (Updated)', index=False)

    def _apply_formatting(self):
        """Apply Excel formatting to all sheets"""
        wb = load_workbook(self.report_path)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Format header row
            self._format_header_row(ws)

            # Auto-fit columns
            self._auto_fit_columns(ws)

            # Freeze top row
            ws.freeze_panes = 'A2'

            # Apply filters
            if ws.max_row > 1:
                ws.auto_filter.ref = ws.dimensions

            # Apply color coding based on sheet
            if sheet_name == 'Invalid IDs':
                self._apply_error_color(ws)
            elif sheet_name == 'Successfully Imported':
                self._apply_success_color(ws)
            elif sheet_name == 'Different Ward' or sheet_name == 'Duplicates':
                self._apply_warning_color(ws)

        wb.save(self.report_path)

    def _format_header_row(self, ws):
        """Format header row with bold font and background color"""
        header_fill = PatternFill(start_color=self.COLORS['header'],
                                  end_color=self.COLORS['header'],
                                  fill_type='solid')
        header_font = Font(bold=True, size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

    def _auto_fit_columns(self, ws):
        """Auto-fit column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            ws.column_dimensions[column_letter].width = adjusted_width

    def _apply_error_color(self, ws):
        """Apply error color to data rows"""
        error_fill = PatternFill(start_color=self.COLORS['error'],
                                end_color=self.COLORS['error'],
                                fill_type='solid')

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.fill = error_fill

    def _apply_success_color(self, ws):
        """Apply success color to data rows"""
        success_fill = PatternFill(start_color=self.COLORS['success'],
                                   end_color=self.COLORS['success'],
                                   fill_type='solid')

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.fill = success_fill

    def _apply_warning_color(self, ws):
        """Apply warning color to data rows"""
        warning_fill = PatternFill(start_color=self.COLORS['warning'],
                                   end_color=self.COLORS['warning'],
                                   fill_type='solid')

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.fill = warning_fill

