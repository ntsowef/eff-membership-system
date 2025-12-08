import requests
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
import logging
import os
import sys

# ==================== EXCEL PATH ====================
EXCEL_FILE_PATH = "VOTER_REG_STATUS/DANNHAUSER_WARD_6-18-11-2025.xlsx"   
# ==========================================================

# ------------------- CONFIG -------------------
API_USERNAME = 'IECWebAPIPartyEFF'
API_PASSWORD = '85316416dc5b498586ed519e670931e9'
TOKEN_URL = 'https://api.elections.org.za/token'
VOTER_ENDPOINT = 'https://api.elections.org.za/api/Voters/IDNumber/'
OUTPUT_DIR = "VOTER_REG_STATUS"
ID_COLUMN_INDEX = 8        # Column H
EXPECTED_WARD_COLUMN = 5   # Column E
MAX_WORKERS = 15

os.makedirs(OUTPUT_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(OUTPUT_DIR, "auto_check.log"), encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)

def get_access_token():
    data = {'grant_type': 'password', 'username': API_USERNAME, 'password': API_PASSWORD}
    try:
        r = requests.post(TOKEN_URL, data=data, timeout=15)
        r.raise_for_status()
        return r.json()['access_token']
    except Exception as e:
        logging.error(f"Token error: {e}")
        sys.exit(1)

def fetch_voter(id_number, token):
    headers = {'Authorization': f'Bearer {token}'}
    try:
        r = requests.get(f"{VOTER_ENDPOINT}{id_number}", headers=headers, timeout=12)
        if r.status_code == 200:
            return r.json()
        elif r.status_code == 404:
            return {"bRegistered": False, "VotingStation": {}}
        else:
            return None
    except:
        return None

def main():
    if not os.path.isfile(EXCEL_FILE_PATH):
        print("ERROR: File not found!")
        print(f"Path: {EXCEL_FILE_PATH}")
        print("Please edit the script and fix EXCEL_FILE_PATH")
        input("Press Enter to exit...")
        sys.exit(1)

    print("Bakkie Connect - IEC Voter Checker (Auto Mode)")
    print(f"Processing: {os.path.basename(EXCEL_FILE_PATH)}")
    print(f"Results → {os.path.abspath(OUTPUT_DIR)}\n")

    wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
    ws = wb.active

    id_list = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=ID_COLUMN_INDEX).value
        if val:
            clean_id = str(val).strip().zfill(13)[:13]
            if clean_id.isdigit():
                id_list.append((clean_id, row))

    if not id_list:
        print("No valid 13-digit IDs found in column H.")
        sys.exit(1)

    print(f"Found {len(id_list)} IDs → Starting IEC lookup...")

    token = get_access_token()

    reg_in_ward = not_reg = diff_ward = deceased = 0
    reg_list = not_reg_list = diff_list = deceased_list = []

    start_time = time.time()

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(fetch_voter, id_num, token): (id_num, row) for id_num, row in id_list}

        for i, future in enumerate(as_completed(futures), 1):
            id_num, row = futures[future]
            data = future.result()
            print(f"\rProgress: {i}/{len(id_list)}", end="")

            if not data:
                continue

            bReg = data.get('bRegistered', False)
            ward = data.get('VotingStation', {}).get('Delimitation', {}).get('WardID')
            vd_num = data.get('VotingStation', {}).get('Delimitation', {}).get('VDNumber', '')
            vs_name = data.get('VotingStation', {}).get('Name', '')
            suburb = data.get('VotingStation', {}).get('Location', {}).get('Suburb', '')
            street = data.get('VotingStation', {}).get('Location', {}).get('Street', '')

            expected_ward = ws.cell(row=row, column=EXPECTED_WARD_COLUMN).value

            id_cell = ws.cell(row=row, column=ID_COLUMN_INDEX)
            status_cell = ws.cell(row=row, column=27)
            vd_cell = ws.cell(row=row, column=26)

            if bReg and ward == expected_ward:
                id_cell.fill = PatternFill(fgColor="00FF00", fill_type="solid")
                status_cell.value = "REGISTERED IN WARD"
                vd_cell.value = vd_num
                reg_in_ward += 1
                reg_list.append(f"Row {row}: {id_num}")

            elif not bReg and ward in (None, 0, ""):
                id_cell.fill = PatternFill(fgColor="fe2400", fill_type="solid")
                status_cell.value = "NOT REGISTERED VOTER"
                status_cell.fill = PatternFill(fgColor="fe2400", fill_type="solid")
                vd_cell.value = "99999999"
                not_reg += 1
                not_reg_list.append(f"Row {row}: {id_num}")

            elif bReg and ward != expected_ward and ward not in (None, 0):
                id_cell.fill = PatternFill(fgColor="3777b2", fill_type="solid")
                status_cell.value = "REGISTERED IN DIFFERENT WARD"
                vd_cell.value = "22222222"
                diff_ward += 1
                diff_list.append(f"Row {row}: {id_num} → Ward {ward}")

            elif vs_name and not bReg:
                id_cell.fill = PatternFill(fgColor="ffff00", fill_type="solid")
                status_cell.value = "DECEASED"
                status_cell.fill = PatternFill(fgColor="ffff00", fill_type="solid")
                vd_cell.value = "11111111"
                deceased += 1
                deceased_list.append(f"Row {row}: {id_num}")

            addr_cell = ws.cell(row=row, column=14)
            if addr_cell.value in (None, "", "N/A", "CENTURION"):
                addr = f"{suburb} {street}".strip()
                if addr and addr != " ":
                    addr_cell.value = addr

    # Save updated original file
    updated_file = os.path.join(OUTPUT_DIR, os.path.basename(EXCEL_FILE_PATH))
    wb.save(updated_file)

    # FIXED: Proper DataFrame creation for summary
    report_file = os.path.join(OUTPUT_DIR, "VOTER_REGISTRATION_STATUS.xlsx")
    with pd.ExcelWriter(report_file, engine='openpyxl') as writer:
        # Summary sheet
        summary_df = pd.DataFrame({
            "Category": [
                "Total Processed",
                "Registered in Correct Ward",
                "Not Registered",
                "Registered Elsewhere",
                "Deceased"
            ],
            "Count": [len(id_list), reg_in_ward, not_reg, diff_ward, deceased]
        })
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # Other sheets
        pd.DataFrame(reg_list, columns=["Registered in Ward"]).to_excel(writer, sheet_name="CorrectWard", index=False)
        pd.DataFrame(not_reg_list, columns=["Not Registered"]).to_excel(writer, sheet_name="NotRegistered", index=False)
        pd.DataFrame(diff_list, columns=["Different Ward"]).to_excel(writer, sheet_name="WrongWard", index=False)
        pd.DataFrame(deceased_list, columns=["Deceased"]).to_excel(writer, sheet_name="Deceased", index=False)

    elapsed = time.time() - start_time
    print(f"\n\nCOMPLETED in {elapsed:.1f} seconds")
    print("="*60)
    print(f"Total Processed         : {len(id_list)}")
    print(f"Registered in Ward      : {reg_in_ward}")
    print(f"Not Registered          : {not_reg}")
    print(f"Wrong Ward              : {diff_ward}")
    print(f"Deceased                : {deceased}")
    print(f"\nUpdated file → {updated_file}")
    print(f"Full report  → {report_file}")
    print("="*60)
    print("All done. You can close this window.")

    # Optional: auto-open folder
    if os.name == 'nt':  # Windows
        os.startfile(os.path.abspath(OUTPUT_DIR))

if _name_ == "_main_":
    main()